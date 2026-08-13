from app.models.models import User, Exam, Question
from tests.conftest import register, login, logout


def test_register_and_login(client):
    r = register(client, "alice", "alice@example.com", "password1", "student")
    assert b"Registration successful" in r.data

    r = login(client, "alice", "password1")
    assert b"Available Exams" in r.data


def test_duplicate_username_rejected(client):
    register(client, "bob", "bob@example.com", "password1", "student")
    r = register(client, "bob", "bob2@example.com", "password1", "student")
    assert b"Username already taken" in r.data


def test_invalid_login_rejected(client):
    register(client, "carol", "carol@example.com", "password1", "student")
    r = login(client, "carol", "wrongpassword")
    assert b"Invalid username or password" in r.data


def test_teacher_can_create_publish_exam_and_student_can_take_it(client, app):
    # Teacher creates and publishes an exam with one question
    register(client, "teach1", "teach1@example.com", "password1", "teacher")
    login(client, "teach1", "password1")

    r = client.post(
        "/teacher/exams/create",
        data={"title": "Math 101", "description": "Basic math", "duration_minutes": "10"},
        follow_redirects=True,
    )
    assert b"Add a Question" in r.data

    with app.app_context():
        exam = Exam.query.filter_by(title="Math 101").first()
        exam_id = exam.id

    client.post(
        f"/teacher/exams/{exam_id}/questions",
        data={
            "question_text": "What is 2+2?",
            "marks": "5",
            "correct_choice": "1",
            "choice_text": ["3", "4", "5", "6"],
        },
        follow_redirects=True,
    )
    r = client.post(f"/teacher/exams/{exam_id}/publish", data={}, follow_redirects=True)
    assert b"published" in r.data.lower()
    logout(client)

    # Student registers, sees the exam, and takes it correctly
    register(client, "stud1", "stud1@example.com", "password1", "student")
    login(client, "stud1", "password1")

    r = client.get("/student/dashboard")
    assert b"Math 101" in r.data

    with app.app_context():
        question = Question.query.filter_by(exam_id=exam_id).first()
        correct_choice_id = question.correct_choice().id

    r = client.post(
        f"/student/exams/{exam_id}/take",
        data={f"question_{question.id}": str(correct_choice_id)},
        follow_redirects=True,
    )
    assert b"5.0 / 5.0" in r.data

    # Re-visiting the exam should redirect to the already-submitted result,
    # not allow a second attempt.
    r = client.get(f"/student/exams/{exam_id}/take", follow_redirects=True)
    assert b"attempt limit" in r.data.lower() or b"already submitted" in r.data.lower()


def test_wrong_answer_scores_zero(client, app):
    register(client, "teach2", "teach2@example.com", "password1", "teacher")
    login(client, "teach2", "password1")
    client.post(
        "/teacher/exams/create",
        data={"title": "Quiz", "description": "", "duration_minutes": "10"},
        follow_redirects=True,
    )
    with app.app_context():
        exam = Exam.query.filter_by(title="Quiz").first()
        exam_id = exam.id
    client.post(
        f"/teacher/exams/{exam_id}/questions",
        data={
            "question_text": "2+2?",
            "marks": "3",
            "correct_choice": "0",
            "choice_text": ["4", "5"],
        },
        follow_redirects=True,
    )
    client.post(f"/teacher/exams/{exam_id}/publish", data={}, follow_redirects=True)
    logout(client)

    register(client, "stud2", "stud2@example.com", "password1", "student")
    login(client, "stud2", "password1")
    with app.app_context():
        question = Question.query.filter_by(exam_id=exam_id).first()
        wrong_choice = [c for c in question.choices if not c.is_correct][0]

    r = client.post(
        f"/student/exams/{exam_id}/take",
        data={f"question_{question.id}": str(wrong_choice.id)},
        follow_redirects=True,
    )
    assert b"0.0 / 3.0" in r.data


def test_admin_can_see_and_disable_users(client, app):
    register(client, "stud3", "stud3@example.com", "password1", "student")
    logout(client)

    login(client, "admin", "Admin@123")
    r = client.get("/admin/dashboard")
    assert b"stud3" in r.data

    with app.app_context():
        target = User.query.filter_by(username="stud3").first()
        target_id = target.id

    r = client.post(f"/admin/users/{target_id}/toggle", data={}, follow_redirects=True)
    assert b"disabled" in r.data.lower()
    logout(client)

    r = login(client, "stud3", "password1")
    assert b"disabled" in r.data.lower()


def test_role_access_is_enforced(client):
    register(client, "stud4", "stud4@example.com", "password1", "student")
    login(client, "stud4", "password1")

    # Student should not be able to reach teacher-only or admin-only routes
    r = client.get("/teacher/dashboard")
    assert r.status_code == 403

    r = client.get("/admin/dashboard")
    assert r.status_code == 403


def test_student_assignment_whitelist(client, app):
    # Register 1 teacher and 2 students
    register(client, "teach_wl", "teach_wl@example.com", "password1", "teacher")
    logout(client)

    register(client, "stud_assigned", "stud_a@example.com", "password1", "student")
    logout(client)

    register(client, "stud_unassigned", "stud_u@example.com", "password1", "student")
    logout(client)

    login(client, "teach_wl", "password1")
    client.post("/teacher/exams/create", data={"title": "Private Exam", "duration_minutes": "15"}, follow_redirects=True)

    with app.app_context():
        exam = Exam.query.filter_by(title="Private Exam").first()
        exam_id = exam.id
        assigned_user = User.query.filter_by(username="stud_assigned").first()
        assigned_id = assigned_user.id

    client.post(
        f"/teacher/exams/{exam_id}/questions",
        data={"question_text": "Sample Q?", "marks": "1", "correct_choice": "0", "choice_text": ["A", "B"]},
        follow_redirects=True,
    )

    # Teacher publishes exam specifically assigned to stud_assigned
    client.post(
        f"/teacher/exams/{exam_id}/publish_settings",
        data={
            "max_attempts": "1",
            "assigned_students": [str(assigned_id)],
            "action": "publish",
        },
        follow_redirects=True,
    )
    logout(client)

    # Assigned student logs in -> sees exam
    login(client, "stud_assigned", "password1")
    r = client.get("/student/dashboard")
    assert b"Private Exam" in r.data
    logout(client)

    # Unassigned student logs in -> exam is NOT listed in dashboard
    login(client, "stud_unassigned", "password1")
    r = client.get("/student/dashboard")
    assert b"Private Exam" not in r.data

    # Direct access attempt by unassigned student is blocked
    r = client.get(f"/student/exams/{exam_id}/take", follow_redirects=True)
    assert b"not assigned" in r.data.lower()


def test_multiple_takes_per_student(client, app):
    register(client, "teach_multi", "teach_multi@example.com", "password1", "teacher")
    login(client, "teach_multi", "password1")
    client.post("/teacher/exams/create", data={"title": "Multi Take Exam", "duration_minutes": "10"}, follow_redirects=True)

    with app.app_context():
        exam = Exam.query.filter_by(title="Multi Take Exam").first()
        exam_id = exam.id

    client.post(
        f"/teacher/exams/{exam_id}/questions",
        data={"question_text": "Capital of France?", "marks": "2", "correct_choice": "1", "choice_text": ["Berlin", "Paris"]},
        follow_redirects=True,
    )

    # Configure max_attempts = 2 and publish
    client.post(
        f"/teacher/exams/{exam_id}/publish_settings",
        data={"max_attempts": "2", "action": "publish"},
        follow_redirects=True,
    )
    logout(client)

    register(client, "stud_retake", "stud_retake@example.com", "password1", "student")
    login(client, "stud_retake", "password1")

    with app.app_context():
        question = Question.query.filter_by(exam_id=exam_id).first()
        correct_choice_id = question.correct_choice().id

    # Attempt 1
    r = client.post(
        f"/student/exams/{exam_id}/take",
        data={f"question_{question.id}": str(correct_choice_id)},
        follow_redirects=True,
    )
    assert b"Attempt 1 of 2" in r.data or b"2.0 / 2.0" in r.data

    # Attempt 2 (allowed since max_attempts = 2)
    r = client.post(
        f"/student/exams/{exam_id}/take",
        data={f"question_{question.id}": str(correct_choice_id)},
        follow_redirects=True,
    )
    assert b"Attempt 2 of 2" in r.data or b"2.0 / 2.0" in r.data

    # Attempt 3 (blocked since max_attempts = 2 reached)
    r = client.get(f"/student/exams/{exam_id}/take", follow_redirects=True)
    assert b"attempt limit" in r.data.lower()


def test_shareable_exam_link(client, app):
    register(client, "teach_link", "teach_link@example.com", "password1", "teacher")
    login(client, "teach_link", "password1")
    client.post("/teacher/exams/create", data={"title": "Shareable Link Exam", "duration_minutes": "10"}, follow_redirects=True)

    with app.app_context():
        exam = Exam.query.filter_by(title="Shareable Link Exam").first()
        exam_id = exam.id
        token = exam.access_token

    client.post(
        f"/teacher/exams/{exam_id}/questions",
        data={"question_text": "Q1?", "marks": "1", "correct_choice": "0", "choice_text": ["A", "B"]},
        follow_redirects=True,
    )
    client.post(f"/teacher/exams/{exam_id}/publish", data={}, follow_redirects=True)
    logout(client)

    # Accessing share link unauthenticated redirects to login
    r = client.get(f"/student/exams/link/{token}", follow_redirects=True)
    assert b"log in" in r.data.lower()

    # Accessing share link authenticated as student redirects to take_exam
    register(client, "stud_link", "stud_link@example.com", "password1", "student")
    login(client, "stud_link", "password1")
    r = client.get(f"/student/exams/link/{token}", follow_redirects=True)
    assert b"Shareable Link Exam" in r.data


def test_class_creation_and_individual_student_addition(client, app):
    register(client, "teach_cls", "teach_cls@example.com", "password1", "teacher")
    login(client, "teach_cls", "password1")

    # Create Class
    r = client.post("/teacher/classes", data={"name": "Class 10A", "description": "Math section"}, follow_redirects=True)
    assert b"Class 10A" in r.data and b"created" in r.data

    with app.app_context():
        from app.models.models import Class
        c = Class.query.filter_by(name="Class 10A").first()
        class_id = c.id

    # Add individual student
    r = client.post(
        f"/teacher/classes/{class_id}",
        data={"username": "student_in_class", "email": "sic@example.com", "password": "Password@123"},
        follow_redirects=True,
    )
    assert b"student_in_class" in r.data

    with app.app_context():
        from app.models.models import Class, User
        c = Class.query.get(class_id)
        u = User.query.filter_by(username="student_in_class").first()
        assert u in c.students


def test_csv_and_excel_bulk_upload(client, app):
    import io
    register(client, "teach_bulk", "teach_bulk@example.com", "password1", "teacher")
    login(client, "teach_bulk", "password1")

    client.post("/teacher/classes", data={"name": "Science 101"}, follow_redirects=True)
    with app.app_context():
        from app.models.models import Class
        c = Class.query.filter_by(name="Science 101").first()
        class_id = c.id

    # Simulate CSV upload
    csv_data = "Username,Email,Password\nbulk_stud1,b1@example.com,Pass@123\nbulk_stud2,b2@example.com,Pass@123\n"
    data = {"excel_file": (io.BytesIO(csv_data.encode("utf-8")), "students.csv")}

    r = client.post(
        f"/teacher/classes/{class_id}/upload_excel",
        data=data,
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert b"Excel Import Complete" in r.data

    with app.app_context():
        from app.models.models import Class, User
        c = Class.query.get(class_id)
        u1 = User.query.filter_by(username="bulk_stud1").first()
        u2 = User.query.filter_by(username="bulk_stud2").first()
        assert u1 is not None and u1 in c.students
        assert u2 is not None and u2 in c.students


def test_class_exam_assignment(client, app):
    register(client, "teach_ce", "teach_ce@example.com", "password1", "teacher")
    login(client, "teach_ce", "password1")

    # Create Class & Add student
    client.post("/teacher/classes", data={"name": "History Class"}, follow_redirects=True)
    with app.app_context():
        from app.models.models import Class
        c = Class.query.filter_by(name="History Class").first()
        class_id = c.id

    client.post(f"/teacher/classes/{class_id}", data={"username": "hist_student", "email": "hist@example.com"}, follow_redirects=True)

    # Register another unassigned student
    register(client, "other_student", "other@example.com", "password1", "student")

    # Create Exam & Assign to Class
    login(client, "teach_ce", "password1")
    client.post("/teacher/exams/create", data={"title": "History Final", "duration_minutes": "30"}, follow_redirects=True)
    with app.app_context():
        exam = Exam.query.filter_by(title="History Final").first()
        exam_id = exam.id

    client.post(f"/teacher/exams/{exam_id}/questions", data={"question_text": "Q?", "marks": "1", "correct_choice": "0", "choice_text": ["A", "B"]}, follow_redirects=True)

    client.post(
        f"/teacher/exams/{exam_id}/publish_settings",
        data={
            "max_attempts": "1",
            "assigned_classes": [str(class_id)],
            "action": "publish",
        },
        follow_redirects=True,
    )
    logout(client)

    # Class student logs in -> sees exam
    login(client, "hist_student", "Student@123")
    r = client.get("/student/dashboard")
    assert b"History Final" in r.data
    logout(client)

    # Other student logs in -> exam NOT visible
    login(client, "other_student", "password1")
    r = client.get("/student/dashboard")
    assert b"History Final" not in r.data


def test_download_student_import_template(client):
    register(client, "teach_tmpl", "teach_tmpl@example.com", "password1", "teacher")
    login(client, "teach_tmpl", "password1")

    r = client.get("/teacher/classes/template/download")
    assert r.status_code == 200
    assert "spreadsheetml.sheet" in r.content_type
    assert r.headers["Content-Disposition"].startswith("attachment;")
    assert "student_import_template.xlsx" in r.headers["Content-Disposition"]


def test_export_class_results_excel(client, app):
    import openpyxl
    import io

    register(client, "teach_exp", "teach_exp@example.com", "password1", "teacher")
    login(client, "teach_exp", "password1")

    # Create class and add student
    client.post("/teacher/classes", data={"name": "Physics 101"}, follow_redirects=True)
    with app.app_context():
        from app.models.models import Class
        c = Class.query.filter_by(name="Physics 101").first()
        class_id = c.id

    client.post(f"/teacher/classes/{class_id}", data={"username": "phys_student", "email": "phys@example.com"}, follow_redirects=True)

    # Create & publish exam
    client.post("/teacher/exams/create", data={"title": "Physics Quiz", "duration_minutes": "15"}, follow_redirects=True)
    with app.app_context():
        exam = Exam.query.filter_by(title="Physics Quiz").first()
        exam_id = exam.id

    client.post(f"/teacher/exams/{exam_id}/questions", data={"question_text": "Speed of light?", "marks": "10", "correct_choice": "0", "choice_text": ["3x10^8 m/s", "100 m/s"]}, follow_redirects=True)
    client.post(f"/teacher/exams/{exam_id}/publish_settings", data={"assigned_classes": [str(class_id)], "action": "publish"}, follow_redirects=True)
    logout(client)

    # Student takes exam
    login(client, "phys_student", "Student@123")
    with app.app_context():
        question = Question.query.filter_by(exam_id=exam_id).first()
        correct_choice_id = question.correct_choice().id

    client.post(f"/student/exams/{exam_id}/take", data={f"question_{question.id}": str(correct_choice_id)}, follow_redirects=True)
    logout(client)

    # Teacher exports class results as Excel
    login(client, "teach_exp", "password1")
    r = client.get(f"/teacher/classes/{class_id}/export_results")
    assert r.status_code == 200
    assert "spreadsheetml.sheet" in r.content_type

    # Verify Excel contents with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == [
        "Student Username", "Student Email", "Exam Title",
        "Score Obtained", "Total Marks", "Percentage (%)",
        "Attempt Number", "Submitted At"
    ]
    assert len(rows) == 2
    assert rows[1][0] == "phys_student"
    assert rows[1][2] == "Physics Quiz"
    assert rows[1][3] == 10.0


def test_export_exam_results_excel(client, app):
    import openpyxl
    import io

    register(client, "teach_ex_export", "teach_ee@example.com", "password1", "teacher")
    login(client, "teach_ex_export", "password1")

    # Create & publish exam
    client.post("/teacher/exams/create", data={"title": "Biology Exam", "duration_minutes": "20"}, follow_redirects=True)
    with app.app_context():
        exam = Exam.query.filter_by(title="Biology Exam").first()
        exam_id = exam.id

    client.post(f"/teacher/exams/{exam_id}/questions", data={"question_text": "What is DNA?", "marks": "5", "correct_choice": "0", "choice_text": ["Genetic Material", "Protein"]}, follow_redirects=True)
    client.post(f"/teacher/exams/{exam_id}/publish", data={}, follow_redirects=True)
    logout(client)

    # Student takes exam
    register(client, "bio_stud", "bio@example.com", "password1", "student")
    login(client, "bio_stud", "password1")
    with app.app_context():
        question = Question.query.filter_by(exam_id=exam_id).first()
        correct_choice_id = question.correct_choice().id

    client.post(f"/student/exams/{exam_id}/take", data={f"question_{question.id}": str(correct_choice_id)}, follow_redirects=True)
    logout(client)

    # Teacher views /exams/<id>/results and exports results to Excel
    login(client, "teach_ex_export", "password1")
    r = client.get(f"/teacher/exams/{exam_id}/results")
    assert b"Download Results (Excel)" in r.data

    r = client.get(f"/teacher/exams/{exam_id}/export_results")
    assert r.status_code == 200
    assert "spreadsheetml.sheet" in r.content_type
    assert "Biology_Exam_Results.xlsx" in r.headers["Content-Disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == [
        "Student Username", "Student Email", "Score Obtained",
        "Total Marks", "Percentage (%)", "Attempt Number", "Submitted At"
    ]
    assert len(rows) == 2
    assert rows[1][0] == "bio_stud"
    assert rows[1][2] == 5.0




