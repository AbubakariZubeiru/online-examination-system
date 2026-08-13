import csv
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, send_file
from flask_login import login_required, current_user

from app import db
from app.models.models import Exam, Question, Choice, Submission, User, Class
from app.utils import role_required

teacher_bp = Blueprint("teacher", __name__, url_prefix="/teacher")


def _get_owned_exam_or_404(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    if exam.teacher_id != current_user.id:
        abort(403)
    return exam


def _get_owned_class_or_404(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.teacher_id != current_user.id:
        abort(403)
    return class_obj


@teacher_bp.route("/dashboard")
@login_required
@role_required("teacher")
def dashboard():
    exams = Exam.query.filter_by(teacher_id=current_user.id).order_by(
        Exam.created_at.desc()
    ).all()
    # Ensure tokens exist for shareable links
    for exam in exams:
        exam.generate_token()
    db.session.commit()
    return render_template("teacher/dashboard.html", exams=exams)


@teacher_bp.route("/classes", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def manage_classes():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()

        if not name:
            flash("Class name is required.", "danger")
        else:
            c = Class(name=name, description=description, teacher_id=current_user.id)
            db.session.add(c)
            db.session.commit()
            flash(f"Class '{c.name}' created successfully!", "success")
            return redirect(url_for("teacher.manage_classes"))

    classes = Class.query.filter_by(teacher_id=current_user.id).order_by(Class.created_at.desc()).all()
    return render_template("teacher/classes.html", classes=classes)


@teacher_bp.route("/classes/<int:class_id>", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def class_detail(class_id):
    class_obj = _get_owned_class_or_404(class_id)

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        errors = []
        if not username or not email:
            errors.append("Username and Email are required.")

        if errors:
            for e in errors:
                flash(e, "danger")
        else:
            user = User.query.filter((User.username == username) | (User.email == email)).first()
            if not user:
                if not password:
                    password = "Student@123"
                user = User(username=username, email=email, role="student")
                user.set_password(password)
                db.session.add(user)
                db.session.commit()
                flash(f"Created student account '{username}'.", "success")

            if user.role != "student":
                flash(f"User '{username}' is an admin/teacher and cannot be added as a student.", "warning")
            elif user in class_obj.students:
                flash(f"Student '{username}' is already in this class.", "info")
            else:
                class_obj.students.append(user)
                db.session.commit()
                flash(f"Added student '{username}' to class '{class_obj.name}'.", "success")

        return redirect(url_for("teacher.class_detail", class_id=class_obj.id))

    return render_template("teacher/class_detail.html", class_obj=class_obj)


@teacher_bp.route("/classes/<int:class_id>/upload_excel", methods=["POST"])
@login_required
@role_required("teacher")
def upload_excel_students(class_id):
    class_obj = _get_owned_class_or_404(class_id)

    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Please select an Excel (.xlsx, .xls) or CSV (.csv) file to upload.", "danger")
        return redirect(url_for("teacher.class_detail", class_id=class_obj.id))

    filename = file.filename.lower()
    imported_count = 0
    added_to_class_count = 0
    skipped_count = 0

    student_rows = []

    try:
        if filename.endswith(".csv"):
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
            csv_reader = csv.DictReader(stream)
            for row in csv_reader:
                norm_row = {k.strip().lower(): v.strip() for k, v in row.items() if k and v}
                student_rows.append(norm_row)
        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(file.stream, data_only=True)
            sheet = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = str(val or "").strip()
                student_rows.append(row_dict)
        else:
            flash("Unsupported file format. Upload .xlsx, .xls, or .csv.", "danger")
            return redirect(url_for("teacher.class_detail", class_id=class_obj.id))
    except Exception as err:
        flash(f"Error reading file: {str(err)}", "danger")
        return redirect(url_for("teacher.class_detail", class_id=class_obj.id))

    for row in student_rows:
        username = row.get("username") or row.get("user") or row.get("student_username")
        email = row.get("email") or row.get("student_email")
        password = row.get("password") or "Student@123"

        if not username or not email:
            skipped_count += 1
            continue

        user = User.query.filter((User.username == username) | (User.email == email)).first()
        if not user:
            user = User(username=username, email=email, role="student")
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            imported_count += 1

        if user.role == "student" and user not in class_obj.students:
            class_obj.students.append(user)
            added_to_class_count += 1

    db.session.commit()
    flash(
        f"Excel Import Complete: Created {imported_count} new account(s), "
        f"added {added_to_class_count} student(s) to '{class_obj.name}' "
        f"({skipped_count} invalid/skipped rows).",
        "success",
    )
    return redirect(url_for("teacher.class_detail", class_id=class_obj.id))


@teacher_bp.route("/classes/<int:class_id>/delete", methods=["POST"])
@login_required
@role_required("teacher")
def delete_class(class_id):
    class_obj = _get_owned_class_or_404(class_id)
    db.session.delete(class_obj)
    db.session.commit()
    flash(f"Class '{class_obj.name}' deleted.", "success")
    return redirect(url_for("teacher.manage_classes"))


@teacher_bp.route("/classes/<int:class_id>/remove_student/<int:student_id>", methods=["POST"])
@login_required
@role_required("teacher")
def remove_student_from_class(class_id, student_id):
    class_obj = _get_owned_class_or_404(class_id)
    student = User.query.get_or_404(student_id)
    if student in class_obj.students:
        class_obj.students.remove(student)
        db.session.commit()
        flash(f"Removed '{student.username}' from class '{class_obj.name}'.", "success")
    return redirect(url_for("teacher.class_detail", class_id=class_obj.id))


@teacher_bp.route("/exams/create", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def create_exam():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        duration = request.form.get("duration_minutes", "30")

        if not title:
            flash("Exam title is required.", "danger")
            return render_template("teacher/create_exam.html", form_data=request.form)

        try:
            duration = int(duration)
            if duration <= 0:
                raise ValueError
        except ValueError:
            flash("Duration must be a positive number of minutes.", "danger")
            return render_template("teacher/create_exam.html", form_data=request.form)

        exam = Exam(
            title=title,
            description=description,
            duration_minutes=duration,
            teacher_id=current_user.id,
        )
        exam.generate_token()
        db.session.add(exam)
        db.session.commit()
        flash("Exam created. Now add some questions to it.", "success")
        return redirect(url_for("teacher.manage_questions", exam_id=exam.id))

    return render_template("teacher/create_exam.html", form_data={})


@teacher_bp.route("/exams/<int:exam_id>/questions", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def manage_questions(exam_id):
    exam = _get_owned_exam_or_404(exam_id)

    if request.method == "POST":
        question_text = request.form.get("question_text", "").strip()
        marks = request.form.get("marks", "1")
        choice_texts = request.form.getlist("choice_text")
        correct_index = request.form.get("correct_choice")

        errors = []
        if not question_text:
            errors.append("Question text is required.")

        choice_texts = [c.strip() for c in choice_texts if c.strip()]
        if len(choice_texts) < 2:
            errors.append("Provide at least 2 answer choices.")
        if correct_index is None or not correct_index.isdigit():
            errors.append("Select which choice is correct.")
        elif int(correct_index) >= len(choice_texts):
            errors.append("Correct choice selection is invalid.")

        try:
            marks = int(marks)
            if marks <= 0:
                raise ValueError
        except ValueError:
            errors.append("Marks must be a positive integer.")
            marks = 1

        if errors:
            for e in errors:
                flash(e, "danger")
            return redirect(url_for("teacher.manage_questions", exam_id=exam.id))

        question = Question(exam_id=exam.id, question_text=question_text, marks=marks)
        db.session.add(question)
        db.session.flush()

        for i, text in enumerate(choice_texts):
            choice = Choice(
                question_id=question.id,
                choice_text=text,
                is_correct=(i == int(correct_index)),
            )
            db.session.add(choice)

        db.session.commit()
        flash("Question added.", "success")
        return redirect(url_for("teacher.manage_questions", exam_id=exam.id))

    return render_template("teacher/manage_questions.html", exam=exam)


@teacher_bp.route("/exams/<int:exam_id>/questions/<int:question_id>/delete", methods=["POST"])
@login_required
@role_required("teacher")
def delete_question(exam_id, question_id):
    exam = _get_owned_exam_or_404(exam_id)
    question = Question.query.filter_by(id=question_id, exam_id=exam.id).first_or_404()
    db.session.delete(question)
    db.session.commit()
    flash("Question deleted.", "success")
    return redirect(url_for("teacher.manage_questions", exam_id=exam.id))


@teacher_bp.route("/exams/<int:exam_id>/publish_settings", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def publish_settings(exam_id):
    exam = _get_owned_exam_or_404(exam_id)
    exam.generate_token()
    db.session.commit()

    all_students = User.query.filter_by(role="student").order_by(User.username).all()
    teacher_classes = Class.query.filter_by(teacher_id=current_user.id).order_by(Class.name).all()

    if request.method == "POST":
        start_time_str = request.form.get("start_time", "").strip()
        end_time_str = request.form.get("end_time", "").strip()
        max_attempts_str = request.form.get("max_attempts", "1").strip()
        assigned_student_ids = request.form.getlist("assigned_students")
        assigned_class_ids = request.form.getlist("assigned_classes")

        start_time = None
        end_time = None
        errors = []

        if start_time_str:
            try:
                start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
            except ValueError:
                errors.append("Invalid start date/time format.")

        if end_time_str:
            try:
                end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")
            except ValueError:
                errors.append("Invalid end date/time format.")

        if start_time and end_time and end_time <= start_time:
            errors.append("End date/time must be after the start date/time.")

        try:
            max_attempts = int(max_attempts_str)
            if max_attempts <= 0:
                raise ValueError
        except ValueError:
            errors.append("Number of allowed takes per student must be at least 1.")
            max_attempts = 1

        if errors:
            for e in errors:
                flash(e, "danger")
            shareable_link = url_for("student.exam_link", token=exam.access_token, _external=True)
            return render_template(
                "teacher/publish_settings.html",
                exam=exam,
                all_students=all_students,
                teacher_classes=teacher_classes,
                shareable_link=shareable_link,
            )

        exam.start_time = start_time
        exam.end_time = end_time
        exam.max_attempts = max_attempts

        # Update assigned individual students
        if assigned_student_ids:
            selected_users = User.query.filter(
                User.id.in_([int(sid) for sid in assigned_student_ids if sid.isdigit()])
            ).all()
            exam.assigned_students = selected_users
        else:
            exam.assigned_students = []

        # Update assigned classes
        if assigned_class_ids:
            selected_classes = Class.query.filter(
                Class.id.in_([int(cid) for cid in assigned_class_ids if cid.isdigit()])
            ).all()
            exam.assigned_classes = selected_classes
        else:
            exam.assigned_classes = []

        action = request.form.get("action", "save")
        if action == "publish":
            if not exam.questions:
                flash("Add at least one question before publishing.", "warning")
                db.session.commit()
                return redirect(url_for("teacher.manage_questions", exam_id=exam.id))
            exam.is_published = True
            flash(f"Exam '{exam.title}' publishing settings saved and published!", "success")
        elif action == "unpublish":
            exam.is_published = False
            flash(f"Exam '{exam.title}' unpublished.", "info")
        else:
            flash(f"Exam settings for '{exam.title}' updated successfully.", "success")

        db.session.commit()
        return redirect(url_for("teacher.dashboard"))

    shareable_link = url_for("student.exam_link", token=exam.access_token, _external=True)
    return render_template(
        "teacher/publish_settings.html",
        exam=exam,
        all_students=all_students,
        teacher_classes=teacher_classes,
        shareable_link=shareable_link,
    )


@teacher_bp.route("/exams/<int:exam_id>/publish", methods=["POST"])
@login_required
@role_required("teacher")
def publish_exam(exam_id):
    exam = _get_owned_exam_or_404(exam_id)
    exam.generate_token()
    if not exam.questions:
        flash("Add at least one question before publishing.", "warning")
        return redirect(url_for("teacher.manage_questions", exam_id=exam.id))

    exam.is_published = not exam.is_published
    db.session.commit()
    state = "published" if exam.is_published else "unpublished"
    flash(f"Exam '{exam.title}' has been {state}.", "success")
    return redirect(url_for("teacher.dashboard"))


@teacher_bp.route("/exams/<int:exam_id>/delete", methods=["POST"])
@login_required
@role_required("teacher")
def delete_exam(exam_id):
    exam = _get_owned_exam_or_404(exam_id)
    db.session.delete(exam)
    db.session.commit()
    flash(f"Exam '{exam.title}' deleted.", "success")
    return redirect(url_for("teacher.dashboard"))


@teacher_bp.route("/exams/<int:exam_id>/results")
@login_required
@role_required("teacher")
def exam_results(exam_id):
    exam = _get_owned_exam_or_404(exam_id)
    submissions = Submission.query.filter_by(exam_id=exam.id).order_by(
        Submission.submitted_at.desc()
    ).all()
    return render_template("teacher/exam_results.html", exam=exam, submissions=submissions)


@teacher_bp.route("/exams/<int:exam_id>/export_results")
@login_required
@role_required("teacher")
def export_exam_results(exam_id):
    """Exports an Excel spreadsheet of student results for a specific exam."""
    exam = _get_owned_exam_or_404(exam_id)
    submissions = (
        Submission.query.filter_by(exam_id=exam.id)
        .order_by(Submission.submitted_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{exam.title[:25]} Results"

    headers = [
        "Student Username",
        "Student Email",
        "Score Obtained",
        "Total Marks",
        "Percentage (%)",
        "Attempt Number",
        "Submitted At",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for sub in submissions:
        percentage = round((sub.score / sub.total_marks * 100), 2) if sub.total_marks > 0 else 0.0
        submitted_str = sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else "N/A"

        ws.append([
            sub.student.username,
            sub.student.email,
            round(sub.score, 2),
            round(sub.total_marks, 2),
            f"{percentage}%",
            sub.attempt_number or 1,
            submitted_str,
        ])

    # Auto-adjust column widths
    column_widths = [20, 28, 15, 15, 16, 16, 22]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{exam.title.replace(' ', '_')}_Results.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@teacher_bp.route("/classes/template/download")
@login_required
@role_required("teacher")
def download_student_template():
    """Generates and downloads a sample Excel template for importing student rosters."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Roster Template"

    headers = ["Username", "Email", "Password"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add sample placeholder data
    sample_data = [
        ["john_doe", "john@example.com", "Student@123"],
        ["jane_smith", "jane@example.com", "Student@123"],
        ["alex_brown", "alex@example.com", "Student@123"],
    ]
    for row in sample_data:
        ws.append(row)

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="student_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@teacher_bp.route("/classes/<int:class_id>/export_results")
@login_required
@role_required("teacher")
def export_class_results(class_id):
    """Exports an Excel spreadsheet of exam results for all students in a specific class."""
    class_obj = _get_owned_class_or_404(class_id)
    students = class_obj.students
    student_ids = [s.id for s in students]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name[:25]} Results"

    headers = [
        "Student Username",
        "Student Email",
        "Exam Title",
        "Score Obtained",
        "Total Marks",
        "Percentage (%)",
        "Attempt Number",
        "Submitted At",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if student_ids:
        submissions = (
            Submission.query.filter(Submission.student_id.in_(student_ids))
            .order_by(Submission.submitted_at.desc())
            .all()
        )

        for sub in submissions:
            percentage = round((sub.score / sub.total_marks * 100), 2) if sub.total_marks > 0 else 0.0
            submitted_str = sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else "N/A"

            ws.append([
                sub.student.username,
                sub.student.email,
                sub.exam.title,
                round(sub.score, 2),
                round(sub.total_marks, 2),
                f"{percentage}%",
                sub.attempt_number or 1,
                submitted_str,
            ])

    # Auto-adjust column widths
    column_widths = [20, 28, 30, 15, 15, 16, 16, 22]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{class_obj.name.replace(' ', '_')}_Exam_Results.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


